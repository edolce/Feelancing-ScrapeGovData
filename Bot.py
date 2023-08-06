import io
import queue
import threading
import time
from concurrent.futures.thread import ThreadPoolExecutor
import openpyxl
import requests
import base64
import xlsxwriter as xlsxwriter
import pytesseract
from PIL import Image
import easyocr

pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files/Tesseract-OCR/tesseract.exe'
MAX_WORKERS = 150
reader = easyocr.Reader(['en'])

def decode(data):
    decoded_data = base64.b64decode(data)
    image = Image.open(io.BytesIO(decoded_data))
    text = pytesseract.image_to_string(image, config='--psm 6')
    return text

def easyocr(data):
    decoded_data = base64.b64decode(data)
    image = Image.open(io.BytesIO(decoded_data))
    result = reader.readtext(image)
    return result[0][1]

def get_column_data(file_path, column_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    column_data = []

    column_index = None
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == column_name:
                column_index = cell.column_letter
                break
        if column_index:
            break

    if column_index:
        for cell in sheet[column_index]:
            if cell.row != 1:
                column_data.append(cell.value)

    workbook.close()
    return column_data


class ElectoralSearchBot:
    def __init__(self):
        self.result_queue = queue.Queue()
        self.WrongCaptcha = 0
        self.WrongCaptchaDecoded = 0
        self.MainTimeOut = 0
        self.Solved = 0
        self.Crashes = 0
        self.start_time = time.time()
        self.session = requests.Session()
        self.session.headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json',
            'Origin': 'https://electoralsearch.eci.gov.in',
            'Referer': 'https://electoralsearch.eci.gov.in/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 OPR/99.0.0.0',
            'applicationName': 'ELECTORAL_SEARCH',
            'sec-ch-ua': '"Opera";v="99", "Chromium";v="113", "Not-A.Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
        }

    def single_request(self, epic_number):
        retry_chance = 5
        while True:
            try:
                response = self.session.get("https://gateway.eci.gov.in/api/v1/captcha-service/generateCaptcha", timeout=1)
            except Exception as e:
                self.WrongCaptcha += 1
                continue

            base64_data = response.json()["captcha"]
            captcha_id = response.json()["id"]
            captchaData = easyocr(base64_data)

            json_data = {
                "captchaData": captchaData,
                "captchaId": captcha_id,
                "epicNumber": epic_number,
                "isPortal": True,
                "securityKey": "na"
            }

            try:
                response2 = self.session.post('https://gateway.eci.gov.in/api/v1/elastic/search-by-epic-from-national-display',
                                              json=json_data, timeout=1)
            except Exception as e:
                self.MainTimeOut += 1
                continue

            if response2.status_code == 200:
                data = response2.json()
                return data

            self.WrongCaptchaDecoded += 1
            if retry_chance == 0:
                data = {"error": "retry limit exceeded"}
                return data

            retry_chance -= 1

    def worker(self, epic_number):
        try:
            result = self.single_request(epic_number)
            self.result_queue.put((epic_number, result))
            print("\n\n\n\n")
            self.Solved += 1
            print("Solved: %s" % self.Solved)
            current_time = time.time()
            time_elapsed = current_time - self.start_time
            rate = self.Solved / time_elapsed
            print("Rate: %s" % rate)
            print("Crashes: %s" % self.Crashes)
            print("TimedOUT Captcha requests: %s" % self.WrongCaptcha)
            print("Wrong Captchas: %s" % self.WrongCaptchaDecoded)
            print("Main request timeouts: %s" % self.MainTimeOut)
            print("Alive threads: %s" % threading.active_count())
        except Exception as e:
            print(e)
            self.Crashes += 1

    def process_data(self):
        print("Starting...")

        epic_numbers = get_column_data('data.xlsx', 'UID')
        print("Loaded %s epic numbers" % len(epic_numbers))

        # Using ThreadPoolExecutor with a reduced number of workers for better CPU utilization
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for epic_number in epic_numbers:
                executor.submit(self.worker, epic_number)

        isFinished = True
        print("All threads finished their work.")
        end_time = time.time()
        finalData = {}
        while not self.result_queue.empty():
            epic_number, data = self.result_queue.get()
            finalData[epic_number] = data

        for epic_number, data in finalData.items():
            print(epic_number, data)

        total_time = end_time - self.start_time
        print("Total number of requests:", len(epic_numbers))
        print("Total time taken:", total_time, "seconds")
        print("Average request per second:", len(epic_numbers)/total_time)
        print("success rate:", self.Solved/self.WrongCaptchaDecoded)

        workbook = xlsxwriter.Workbook('results.xlsx')
        worksheet = workbook.add_worksheet()
        row = 0
        col = 0
        value_list = ["S. NO.", "Epic Number", "Name", "Name1", "Age", "Relative Name", "Relative Name1", "State",
                      "District", "Assembly Constituency", "Part", "Polling Station", "Part Serial Number"]
        for i in range(len(value_list)):
            worksheet.write(row, col + i, value_list[i])

        for epic_number, data in finalData.items():
            if not data:
                row += 1
                worksheet.write(row, col, row)
                worksheet.write(row, col + 1, epic_number)
                worksheet.write(row, col + 2, "No data")
                continue
            if "error" in data:
                row += 1
                worksheet.write(row, col, row)
                worksheet.write(row, col + 1, epic_number)
                worksheet.write(row, col + 2, data["error"])
                continue
            for i in range(len(data)):
                row += 1
                content = data[i]["content"]
                NO = row
                epicNumber = content["epicNumber"]
                applicantFirstName = content["applicantFirstName"]
                applicantFirstNameL1 = content["applicantFirstNameL1"]
                age = content["age"]
                relationName = content["relationName"]
                relationNameL1 = content["relationNameL1"]
                stateName = content["stateName"]
                district = str(content["districtNo"]) + "-" + str(content["districtValue"])
                assemble = str(content["acNumber"]) + "-" + str(content["asmblyName"])
                Part = str(content["partNumber"]) + "-" + str(content["partName"])
                psbuildingName = content["psbuildingName"]
                partSerialNumber = content["partSerialNumber"]
                value_list = []
                value_list.append(
                    [NO, epicNumber, applicantFirstName, applicantFirstNameL1, age, relationName, relationNameL1,
                     stateName, district, assemble, Part, psbuildingName, partSerialNumber])

                for value in value_list:
                    for j in range(len(value)):
                        worksheet.write(row, col + j, value[j])

        workbook.close()

if __name__ == '__main__':
    bot = ElectoralSearchBot()
    bot.process_data()