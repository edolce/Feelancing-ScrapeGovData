import concurrent
import io
import json
import os
import queue
import signal
import threading
import time
from concurrent.futures import wait, FIRST_COMPLETED
from concurrent.futures.thread import ThreadPoolExecutor

import openpyxl
import requests
import base64
import xlsxwriter as xlsxwriter
from threading import Lock
import pytesseract
from PIL import Image

critical_function_lock = Lock()
pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files/Tesseract-OCR/tesseract.exe'
Threads ={}
def decode(data):
    # decode base64 string data
    decoded_data = base64.b64decode(data)
    # img_file = open('image.jpg', 'wb')
    # img_file.write(decoded_data)
    # img_file.close()
    #
    # img = cv2.imread('image.jpg')
    image = Image.open(io.BytesIO(decoded_data))
    #save image
    text = pytesseract.image_to_string(image, config='--psm 6')
    return text
WrongCaptcha = 0
WrongCaptchaDecoded = 0
MainTimeOut = 0


def single_request(epicNumber):
    global WrongCaptcha, WrongCaptchaDecoded, MainTimeOut
    retryChance = 5
    while True:
        headers0 = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Origin': 'https://electoralsearch.eci.gov.in',
            'Referer': 'https://electoralsearch.eci.gov.in/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 OPR/99.0.0.0',
            'appName': 'ELECTORAL_SEARCH',
            'sec-ch-ua': '"Opera";v="99", "Chromium";v="113", "Not-A.Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
        }
        # print("[%s] requesting captcha..." % epicNumber)
        try:
            response = requests.get("https://gateway.eci.gov.in/api/v1/captcha-service/generateCaptcha",
                                    headers=headers0, timeout=1)

        except Exception as e:
            #print(e)
            #print("[%s] captcha request timed out" % epicNumber)
            WrongCaptcha +=1
            # print("[%s] retrying captcha..." % epicNumber)
            continue
        # base64 = result[0]["captcha"]
        # captchaId = result[0]["id"]

        base64 = response.json()["captcha"]
        captchaId = response.json()["id"]

        captcha_id = decode(base64)

        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json',
            'Origin': 'https://electoralsearch.eci.gov.in',
            'Referer': 'https://electoralsearch.eci.gov.in/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 OPR/99.0.0.0',
            'applicationName': 'ELECTORAL_SEARCH',
            'sec-ch-ua': '"Opera";v="99", "Chromium";v="113", "Not-A.Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
        }

        json_data = {
            "captchaData": captcha_id,
            "captchaId": captchaId,
            "epicNumber": epicNumber,
            "isPortal": True,
            "securityKey": "na"
        }
        try:
            response2 = requests.post(
                'https://gateway.eci.gov.in/api/v1/elastic/search-by-epic-from-national-display',
                headers=headers,
                json=json_data,
                # cookies=responseCookie[0]
                cookies=response.cookies.get_dict(),
                timeout=1
            )
        except Exception as e:
            #print(e)
            #print("[%s] request MAIN timed out" % epicNumber)
            MainTimeOut += 1
            continue
        # if response2.status_code != 200 repeat the request
        if response2.status_code == 200:
            #print("\033[32m[%s] success\033[0m" % epicNumber)
            data = response2.json()
            return data
        #print("[%s] error captcha is wrong" % epicNumber)
        WrongCaptchaDecoded += 1
        if retryChance == 0:
            #print("[%s] retry limit exceeded" % epicNumber)
            data = {"error": "retry limit exceeded"}
            return data
        # wait 5 seconds before retrying
        # print("[%s] retrying..." % epicNumber)
        retryChance -= 1

result_queue = queue.Queue()
Solved = 0
Crashes = 0

start_time = time.time()
def worker(epic_number):

    global Solved,Crashes

    try:
        # adding this thread to the list
        result = single_request(epic_number)
        result_queue.put((epic_number, result))
        #clear console
        print("\n\n\n\n")
        #print result
        Solved += 1
        print("Solved: %s" % Solved)
        #calculate current rate
        current_time = time.time()
        time_elapsed = current_time - start_time
        rate = Solved / time_elapsed
        print("Rate: %s" % rate)
        # Current crashes
        print("Crashes: %s" % Crashes)
        # Current wrong captchas
        print("TimedOUT Captcha reuests: %s" % WrongCaptcha)
        # Current wrong captchas decoded
        print("Wrong Captchas: %s" % WrongCaptchaDecoded)
        # Current Main request timeouts
        print("Main request timeouts: %s" % MainTimeOut)
        # alive threads
        print("Alive threads: %s" % threading.active_count())

    except Exception as e:
        print(e)
        # Handle the exception here, for example, log the error.
        print(f"Thread crashed: %s" % epic_number)
        Crashes += 1

def get_column_data(file_path, column_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    column_data = []

    # Get the column index based on the column name
    column_index = None
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == column_name:
                column_index = cell.column_letter
                break
        if column_index:
            break

    # If the column is found, fetch its data
    if column_index:
        for cell in sheet[column_index]:
            if cell.row != 1:  # Skip the header row
                column_data.append(cell.value)

    workbook.close()

    return column_data

if __name__ == '__main__':
    print("Starting...")


    # read epic numbers from file
    epic_numbers = get_column_data('data.xlsx', 'UID')
    print("Loaded %s epic numbers" % len(epic_numbers))


    # for epic_number in epic_numbers:
    #     print("Processing epic number: " + epic_number)
    #     finalData.append(single_request(epic_number))

    # Start all the threads.

    print("Starting all threads...")
    # threads = []
    # for epic_number in epic_numbers:
    #     thread = threading.Thread(target=process_epic_number, args=(epic_number,))
    #     threads.append(thread)
    #     Threads[epic_number] = thread


    # Create a ThreadPoolExecutor with the maximum number of concurrent threads
    #divide epic numbers into chunks
    epic_numbers = [epic_numbers[i:i + 500] for i in range(0, len(epic_numbers), 500)]
    for epic_numbers_chunk in epic_numbers:
        with ThreadPoolExecutor(max_workers=50) as executor:
            future_to_epic_number = [executor.submit(worker, epic_number) for epic_number in epic_numbers_chunk]

    # Submit tasks to the thread pool
        # for epic_number in epic_numbers:
        #     executor.submit(process_epic_number, epic_number)
    # with ThreadPool(150) as pool:
    #     # Start the threads using map()
    #     pool.map(lambda thread: thread.run(), threads)

    isFinished = True
    print("All threads finished their work.")
    end_time = time.time()
    # Retrieve the results from the queue
    finalData = {}
    while not result_queue.empty():
        epic_number, data = result_queue.get()
        finalData[epic_number] = data

    # Print the results of the `single_request()` function.
    for epic_number, data in finalData.items():
        print(epic_number, data)

    total_time = end_time - start_time
    print("Total number of requests:", len(epic_numbers))
    print("Total time taken:", total_time, "seconds")
    print("Average time per request:", total_time / len(epic_numbers), "seconds")

    # Write the results to an excel file.
    workbook = xlsxwriter.Workbook('results.xlsx')
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0
    value_list = ["S. NO.", "Epic Number", "Name", "Name1", "Age", "Relative Name", "Relative Name1", "State",
                  "District", "Assembly Constituency", "Part", "Polling Station", "Part Serial Number"]
    for i in range(len(value_list)):
        worksheet.write(row, col + i, value_list[i])

    for epic_number, data in finalData.items():
        # check if data is empty
        if not data:
            row += 1
            worksheet.write(row, col, row)
            worksheet.write(row, col + 1, epic_number)
            worksheet.write(row, col + 2, "No data")
            continue
        # check error
        if "error" in data:
            row += 1
            worksheet.write(row, col, row)
            worksheet.write(row, col + 1, epic_number)
            worksheet.write(row, col + 2, data["error"])
            continue
        for i in range(len(data)):
            row += 1
            content = data[i]["content"]
            NO = row
            epicNumber = content["epicNumber"]
            applicantFirstName = content["applicantFirstName"]
            applicantFirstNameL1 = content["applicantFirstNameL1"]
            age = content["age"]
            relationName = content["relationName"]
            relationNameL1 = content["relationNameL1"]
            stateName = content["stateName"]
            district = str(content["districtNo"]) + "-" + str(content["districtValue"])
            assemble = str(content["acNumber"]) + "-" + str(content["asmblyName"])
            Part = str(content["partNumber"]) + "-" + str(content["partName"])
            psbuildingName = content["psbuildingName"]
            partSerialNumber = content["partSerialNumber"]
            value_list = []
            value_list.append(
                [NO, epicNumber, applicantFirstName, applicantFirstNameL1, age, relationName, relationNameL1,
                 stateName, district, assemble, Part, psbuildingName, partSerialNumber])

            for value in value_list:
                for j in range(len(value)):
                    worksheet.write(row, col + j, value[j])

    workbook.close()
